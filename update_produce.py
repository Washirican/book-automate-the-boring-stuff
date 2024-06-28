# !/usr/bin/env python3
import openpyxl
import logging

# Configure the root logger
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Create a logger instance
logger = logging.getLogger(__name__)

logger.debug('Opening the file...')
wb = openpyxl.load_workbook('automate_online-materials\\produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

logger.debug('Searching every row of file...')
for row in range(2, sheet.max_row + 1):
    produce_name = sheet.cell(row=row, column=1).value

    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row, column=2).value = PRICE_UPDATES[produce_name]

        logger.debug(f'Match found. Updating price of {produce_name} to {PRICE_UPDATES[produce_name]}')

logger.debug('Saving a copy of the file...')
wb.save('automate_online-materials\\updatedProduceSales.xlsx')