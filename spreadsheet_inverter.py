# !/usr/bin/env python3
import openpyxl

wb_source = openpyxl.load_workbook('automate_online-materials\\produceSales.xlsx')
sheet_source = wb_source.active

wb_dest = openpyxl.Workbook()
sheet_dest_orig = wb_dest.active
sheet_dest_orig.title = 'Original Data'

# Copy 10 rows from source spreadsheet to destination
n = 10
for i in range(1, n + 1):
    wb_dest['Original Data'].cell(row=i, column=1).value = sheet_source.cell(row=i, column=1).value
    wb_dest['Original Data'].cell(row=i, column=2).value = sheet_source.cell(row=i, column=2).value

# Create new sheet in destination workbook
wb_dest.create_sheet('Transposed Data')

# Transpose and copy data from sheet 1 to sheet 2
max_row = wb_dest['Original Data'].max_row 
max_col = wb_dest['Original Data'].max_column

for i in range(1, max_row + 1):
    for j in range(1, max_col + 1):
        wb_dest['Transposed Data'].cell(row=j, column=i).value = wb_dest['Original Data'].cell(row=i, column=j).value
        # wb_dest['Transposed Data'].cell(row=2, column=i).value = wb_dest['Original Data'].cell(row=i, column=2).value

wb_dest.active = wb_dest['Transposed Data']
wb_dest.save('automate_online-materials\\my_spreadsheet_inverter.xlsx')