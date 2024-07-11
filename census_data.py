# !/usr/bin/env python3
# Tabulates population and number of census tracts for
# each county.
import openpyxl
import pprint
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('automate_online-materials\\censuspopdata.xlsx')

sheet = wb['Population by Census Tract']

county_data = {}
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    # Make sure the key for this state exists.
    county_data.setdefault(state, {})

    # Make sure the key for this county in this state exists.
    county_data[state].setdefault(county, {'tracts': 0, 'pop': 0})

    # Each row represents one census tract, so increment by one.
    county_data[state][county]['tracts'] += 1
    # Increase the county pop by the pop in this census tract.
    county_data[state][county]['pop'] += int(pop)

# Open a new text file and write the contents of countyData to it.
print('Writing results...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(county_data))
resultFile.close()
print('Done.')