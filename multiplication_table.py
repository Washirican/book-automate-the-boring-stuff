# !/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active

bold_font = Font(bold=True)

n = int(input('Enter size of multiplication table (N): '))

for i in range(1, n + 1):
    sheet.cell(row=i + 1, column=1).value = i
    sheet.cell(row=i + 1, column=1).font = bold_font
    sheet.cell(row=1, column=i + 1).value = i
    sheet.cell(row=1, column=i + 1).font = bold_font


for i in range(1, n + 1):
    for j in range(1, n + 1):
        sheet.cell(row=i+1, column=j+1).value = i * j


wb.save('automate_online-materials\\my_multiplication_table.xlsx')
